from openpyxl import Workbook, load_workbook

class CBasicXlsData:

    def __init__(self,fields,type):
        self._fieldNameList = []
        self._fieldData = {}
        if(str(type).__eq__("FIELDNAME")):
            for field in fields:
                self._fieldNameList.append(field)
        elif(str(type).__eq__("FIELDWITHDATA")):
            for key in fields.keys():
                self._fieldNameList.append(key)
            self._fieldData.update(fields)
        else:
            print("[ERROR:][CBasicXlsData:__init__] NOT MATCH TYPE ")

    def updateByData(self,data):
        if(data.__len__() != self._fieldNameList.__len__()):
            print("[ERROR:][CBasicXlsData:insertData] Insert Data length don't match FieldName ")
            return
        else:
            for i in range(data.__len__()):
                self._fieldData[self._fieldNameList[i]] = data[i]

    def updateByDict(self,dictDatas):
        for (k,v) in dictDatas.items():
            self._fieldData.setdefault(k,v)

    def getData(self):
        return self._fieldData

    def getFieldName(self):
        return self._fieldNameList



class CBasicXlsSheet:

    def __init__(self,filename,sheetname="Sheet1",sheetindex=0,type="byindex"):
        self._sheetData = []
        self._filename = filename
        self._sheetname = sheetname
        self._wb = load_workbook(filename)
        if(type.__eq__("byindex")):
            self._ws = self._wb.worksheets[sheetindex]
        else:
            self._ws = self._wb.get_sheet_by_name(sheetname)
        self._row_count = self._ws.max_row
        self._column_count = self._ws.max_column
        print(self._row_count)
        print(self._column_count)
        self._header = []
        for col in range(1,self._column_count):                         #header
            self._header.append(self._ws.cell(row=1,column=col).value)

        for r in range(1,self._row_count):                              #data
            xlsdata = CBasicXlsData(self._header,"FIELDNAME")
            data = []
            for c in range(1,self._column_count):
                data.append(self._ws.cell(row=r+1,column=c).value)
            xlsdata.updateByData(data)
            self._sheetData.append(xlsdata)

    def getHeader(self):
        return self._header

    def getData(self,rownum=0):
        if(rownum == 0 or rownum > self._row_count):
            rownum = self._row_count                  #no header
        for i in range(rownum):
            return self._sheetData[:rownum]


